from flask import Blueprint, jsonify, current_app,request,send_file
import openpyxl
from openpyxl import Workbook
import os
from filelock import FileLock
res_bp = Blueprint('res_collection', __name__)

@res_bp.route('/delete-feedback', methods=['POST'])
def delete_feedback():
    """根据文件名和问题描述删除一条反馈"""
    data = request.json
    filename = data.get('filename')
    problem = data.get('problem')
    if problem is None or str(problem).lower() == 'null':
        problem = ''
    else:
        problem = str(problem).strip()
    feedback_file = os.path.join(current_app.config['DATA_FOLDER'], 'feedback.xlsx')
    if not os.path.exists(feedback_file):
        return jsonify({"success": False, "msg": "文件不存在"})

    wb = openpyxl.load_workbook(feedback_file)
    ws = wb.active
    found = False
    for row in ws.iter_rows(min_row=2):
        row_filename = (str(row[0].value or '')).strip()
        row_problem = (str(row[1].value or '')).strip()
        print(f"Row: {row_problem}, {problem}")
        if row_filename == filename and row_problem == problem:
            ws.delete_rows(row[0].row, 1)
            found = True
            break
    wb.save(feedback_file)
    return jsonify({"success": found})

@res_bp.route('/get-feedback-entries', methods=['GET'])
def get_feedback_entries():
    """返回 feedback.xlsx 的全部内容（JSON 格式）"""
    feedback_file = os.path.join(current_app.config['DATA_FOLDER'], 'feedback.xlsx')
    if not os.path.exists(feedback_file):
        return jsonify({"entries": []})  # 如果没有则返回空

    wb = openpyxl.load_workbook(feedback_file)
    ws = wb.active
    entries = []
    # 跳过标题行（第一行），从第二行开始
    for row in ws.iter_rows(min_row=2, values_only=True):
        filename, problem = row
        entries.append({"filename": filename, "problem": problem})
    return jsonify({"entries": entries})

@res_bp.route('/download-feedback', methods=['GET'])
def download_feedback():
    """提供 feedback.xlsx 的下载，导出时同一文件名的所有问题合并为一行，未查看的文件注明"""
    feedback_file = os.path.join(current_app.config['DATA_FOLDER'], 'feedback.xlsx')
    wb = openpyxl.load_workbook(feedback_file)
    ws = wb.active
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(['文件名', '问题描述'])

    # 收集所有反馈问题
    file_problems = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        filename, problem = row
        if filename is None:
            continue
        if filename not in file_problems:
            file_problems[filename] = []
        if problem:
            file_problems[filename].append(str(problem))

    # 获取True1文件夹下所有json文件名
    true1_folder = os.path.join(current_app.config['DATA_FOLDER'], 'True1')
    all_json_files = [f for f in os.listdir(true1_folder) if f.endswith('.json')]

    # 合并并写入新表
    for filename in all_json_files:
        if filename in file_problems:
            worksheet.append([filename, "\n".join(file_problems[filename])])
        else:
            worksheet.append([filename, "未查看"])

    # 保存到临时文件
    temp_file = os.path.join(current_app.config['DATA_FOLDER'], 'feedback_temp.xlsx')
    workbook.save(temp_file)
    if os.path.exists(feedback_file):
        return send_file(temp_file, as_attachment=True)
    return jsonify({"error": "尚未生成反馈文件"}), 404

# @res_bp.route('/feedback', methods=['POST'])
# def feedback():
#     data = request.json
#     print(data)
#     filename = data.get('filename')
#     problem = data.get('problem')
#     feedback_file = os.path.join(current_app.config['DATA_FOLDER'], 'feedback.xlsx')
#     lock_file = feedback_file + '.lock'
#     with FileLock(lock_file):
#         if not os.path.exists(feedback_file):
#             wb = Workbook()
#             ws = wb.active
#             ws.append(['文件名', '问题描述'])
#             wb.save(feedback_file)
#         wb = openpyxl.load_workbook(feedback_file)
#         ws = wb.active
#         # 检查是否已存在相同的反馈
#         exists = False
#         for row in ws.iter_rows(min_row=2, values_only=True):
#             if str(row[0]) == str(filename) and str(row[1]) == str(problem):
#                 exists = True
#                 break
#         if exists:
#             return jsonify({"message": "该反馈已存在，无需重复提交"})
#         ws.append([filename, problem])
#         wb.save(feedback_file)
#     return jsonify({"message": "反馈已提交"})

@res_bp.route('/feedback', methods=['POST'])
def feedback():
    data = request.json
    print(data)
    filename = data.get('filename')
    problem = data.get('problem')
    feedback_file = os.path.join(current_app.config['DATA_FOLDER'], 'feedback.xlsx')
    lock_file = feedback_file + '.lock'
    with FileLock(lock_file):
        if not os.path.exists(feedback_file):
            wb = Workbook()
            ws = wb.active
            ws.append(['文件名', '问题描述'])
            wb.save(feedback_file)
        wb = openpyxl.load_workbook(feedback_file)
        ws = wb.active
        # 检查是否已存在相同的反馈
        exists = False
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(filename) and str(row[1]) == str(problem):
                exists = True
                break
        # 检查并替换“无问题”
        replaced = False
        for row in ws.iter_rows(min_row=2):  # 不加 values_only
            if str(row[0].value) == str(filename) and str(row[1].value) == "无问题":
                row[1].value = problem
                replaced = True
                break
        if exists:
            return jsonify({"message": "该反馈已存在，无需重复提交"})
        if not replaced:
            ws.append([filename, problem])
        wb.save(feedback_file)
    return jsonify({"message": "反馈已提交"})
@res_bp.route('/get-problems/<filename>', methods=['GET'])
def get_problems(filename):
    """
    根据文件名返回该文件所有反馈问题（只返回问题描述列表）
    """
    feedback_file = os.path.join(current_app.config['DATA_FOLDER'], 'feedback.xlsx')
    if not os.path.exists(feedback_file):
        return jsonify([])

    wb = openpyxl.load_workbook(feedback_file)
    ws = wb.active
    problems = []
    #print('前端请求的 filename:', repr(filename))
    # 跳过标题行
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_filename, problem = row
        if str(row_filename) == str(filename):
            problems.append(problem)
    #print('最终返回:', problems)
    return jsonify(problems)